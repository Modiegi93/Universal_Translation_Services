#!/usr/bin/python3
"""Document translations views"""
import os
from flask import request, jsonify, abort, send_file
from langdetect import detect
from werkzeug.utils import secure_filename
from datetime import datetime
from docx import Document
from pptx import Presentation
from openpyxl import load_workbook
from datetime import datetime
import requests
from api.v1.views import app_views
from flask import Flask, jsonify, abort, request
from translator import storage
from translator.document import DocumentTranslation
from translator.translation_model import TranslationModel, Base

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = '/home/ubuntu/Universal_Translation_Services/uploads'


RAPIDAPI_API_KEY = 'f94d683e50mshe579b860a6a88cap1ebc29jsna3f5e765683f'
RAPIDAPI_HOST = 'google-translate1.p.rapidapi.com'

@app_views.route('/documents', methods=['POST'], strict_slashes=False)
def create_document():
    if not request.files or 'file' not in request.files or 'target_lang' not in request.form:
        abort(400, 'Invalid request')

    file = request.files['file']
    target_lang = request.form['target_lang']

    # Process different file types
    if file.filename.endswith('.docx'):
        document = Document(file)
        input_text = ' '.join([p.text for p in document.paragraphs])
    elif file.filename.endswith('.pptx'):
        presentation = Presentation(file)
        input_text = ' '.join([slide.text for slide in presentation.slides])
    elif file.filename.endswith('.xlsx'):
        workbook = load_workbook(file)
        input_text = ' '.join([cell.value for sheet in workbook for row in sheet.iter_rows() for cell in row if cell.value])
    else:
        input_text = file.read().decode('utf-8')

    headers = {
        'X-RapidAPI-Key': RAPIDAPI_API_KEY,
        'Content-Type': 'application/x-www-form-urlencoded',
        'X-RapidAPI-Host': RAPIDAPI_HOST
    }

    payload = {
        'q': input_text,
        'target': target_lang
    }

    try:
        response = requests.post('https://google-translate1.p.rapidapi.com/language/translate/v2',
                                 headers=headers, data=payload)
        translation = response.json()['data']['translations'][0]['translatedText']

        # Language detection
        source_lang = detect(input_text)

        # Store the document details in the database
        new_document = DocumentTranslation(input_text=file.filename,
                                           source_lang=source_lang,
                                           target_lang=target_lang,
                                           translated_text=translation,
                                           created_at=datetime.utcnow())
        storage.new(new_document)
        storage.save()

        return jsonify({'translated_text': translation, 'source_lang': source_lang}), 200
    except Exception as e:
        abort(500, str(e))

@app_views.route('/documents/<document_id>/download', methods=['GET'],
                 strict_slashes=False)
def download_document(document_id):
    document = storage.get(DocumentTranslation, id=document_id)
    if not document:
        abort(404, 'Document not found')

    # Get the file path of the document
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], document.input_text)

    if os.path.isfile(file_path):
        # Send the file for download
        return send_file(file_path, as_attachment=True,
                         attachment_filename=document.input_text)
    else:
        abort(404, 'File not found')

@app_views.route('/documents', methods=['GET'], strict_slashes=False)
def get_all_documents():
    documents = storage.all(DocumentTranslation)
    document_list = [translation.to_dict() for translation in documents.values()]
    return jsonify({'documents': document_list}), 200

@app_views.route('/documents/<document_id>', methods=['GET'], strict_slashes=False)
def get_document(document_id):
    document = storage.get(DocumentTranslation, document_id)
    if not document:
        abort(404, 'Document translation not found')
    return jsonify({
        'id': document.id,
        'filename': document.input_text,
        'source_lang': document.source_lang,
        'target_lang': document.target_lang,
        'translated_text': document.translated_text,
        'created_at': document.created_at
    }), 200

@app_views.route('/documents/<document_id>', methods=['DELETE'], strict_slashes=False)
def delete_document(document_id):
    document = storage.get(DocumentTranslation, document_id)
    if not document:
        abort(404, 'Document translation not found')
    storage.delete(document)
    storage.save()
    return jsonify({'message': 'Document translation deleted successfully'}), 200
